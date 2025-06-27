from fastapi import APIRouter, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import load_workbook
from huggingface_hub import upload_file
from io import BytesIO
import os
import pandas as pd
from datetime import datetime
from copy import copy

router = APIRouter()

REPO_ID = "doluong6/QuanLy"

# ==== Tạo số báo giá ====
def tao_so_bao_gia():
    file = "counter.txt"
    if os.path.exists(file):
        with open(file, "r") as f:
            so = int(f.read().strip()) + 1
    else:
        so = 47
    with open(file, "w") as f:
        f.write(str(so))
    return f"2025/ĐL6/BG.{so:03d}"

# ==== Sao chép định dạng dòng ====
def sao_chep_dong(ws, src_row, dest_row):
    for col in range(1, ws.max_column + 1):
        c1 = ws.cell(row=src_row, column=col)
        c2 = ws.cell(row=dest_row, column=col)
        c2.value = c1.value
        if c1.has_style:
            c2._style = copy(c1._style)

# ==== Tạo báo giá từ file phiếu yêu cầu ====
@router.post("/upload_pyc")
async def upload_pyc(file: UploadFile = File(...)):
    try:
        filename = file.filename
        with open(f"/tmp/{filename}", "wb") as f:
            f.write(file.file.read())
        path = f"/tmp/{filename}"

        wb = load_workbook(path, data_only=True)
        ws = wb.active

        ten_khach = str(ws["I8"].value or "Khach_hang_khong_ten").strip()
        ma_phieu = str(ws["D3"].value or "khong_ro").strip().replace("/", "-")
        thang = str(ws["BS3"].value or "00")
        nam = str(ws["BX3"].value or "0000")

        try:
            thang = f"{int(thang):02}"
        except:
            thang = "00"

        ten_khach_sach = ten_khach.replace("/", "-").replace("\\", "-")
        thu_muc_khach = f"{ten_khach_sach}_{nam}-{thang}_{ma_phieu}"
        path_in_repo = f"Khach_hang/{thu_muc_khach}/Phieu_yeu_cau/{filename}"

        upload_file(
            path_or_fileobj=path,
            path_in_repo=path_in_repo,
            repo_id=REPO_ID,
            repo_type="space",
            token=os.environ["HF_TOKEN"]
        )

        # ==== Trích thiết bị từ phiếu yêu cầu ====
        data = []
        for row in ws.iter_rows(min_row=24):
            ten_tb = row[3].value
            so_luong = row[64].value
            don_gia = row[67].value
            if not ten_tb:
                break
            data.append({
                "Ten_thiet_bi": str(ten_tb).strip(),
                "Số lượng": float(so_luong or 0),
                "Don_gia": float(don_gia or 0)
            })

        if not data:
            return JSONResponse(content={"error": "Không có thiết bị hợp lệ trong phiếu yêu cầu"}, status_code=400)

        import requests
        response = requests.post(
            url="http://localhost:7860/tao_baogia_tu_danhsach",  # hoặc URL backend tương ứng trên Hugging Face
            data={
                "ten_khach": ten_khach,
                "email": "doluong6@quatest1.com.vn",
                "data": pd.DataFrame(data).to_json(orient="records"),
                "vanchuyen": 0
            }
        )

        if response.status_code != 200:
            return JSONResponse(content={"error": "Không tạo được báo giá tự động"}, status_code=500)

        return JSONResponse(content={"message": f"Đã upload phiếu và tạo báo giá cho {ten_khach}"})

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)
