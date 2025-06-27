
from io import BytesIO
from datetime import datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

def tao_so_bao_gia():
    file = "counter.txt"
    import os
    if os.path.exists(file):
        with open(file, "r") as f:
            so = int(f.read().strip()) + 1
    else:
        so = 47
    with open(file, "w") as f:
        f.write(str(so))
    return f"2025/ĐL6/BG.{so:03d}"

def sao_chep_dong(ws, src_row, dest_row):
    for col in range(1, ws.max_column + 1):
        c1 = ws.cell(row=src_row, column=col)
        c2 = ws.cell(row=dest_row, column=col)
        c2.value = c1.value
        if c1.has_style:
            c2._style = copy(c1._style)

def dien_file_baogia(mau, ten_khach, data, so_bg, vanchuyen):
    wb = load_workbook(mau)
    ws = wb.active
    ws["A4"] = f"Số báo giá: {so_bg}"
    ws["A6"] = f"Kính gửi: {ten_khach}"
    ws["C4"] = f"Hà Nội, ngày {datetime.now().day} tháng {datetime.now().month} năm {datetime.now().year}"
    dong_mau = 9
    so_sp = len(data)
    for _ in range(6): ws.delete_rows(dong_mau)
    ws.insert_rows(dong_mau, so_sp + 5)

    font_sp = Font(name='Times New Roman', size=13)
    font_bold = Font(name='Times New Roman', size=13, bold=True)
    border_all = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for i, (_, row) in enumerate(data.iterrows()):
        r = dong_mau + i
        sao_chep_dong(ws, dong_mau - 1, r)
        ws[f"A{r}"] = i + 1
        ws[f"B{r}"].value = str(row["Ten_thiet_bi"])
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        do_dai = len(ws[f"B{r}"].value)
        ws.row_dimensions[r].height = max(20, 15 * ((do_dai // 30) + 1))
        ws[f"C{r}"] = "Cái"
        ws[f"D{r}"] = row["Số lượng"]
        ws[f"E{r}"] = row["Don_gia"]
        ws[f"E{r}"].number_format = '#,##0'
        ws[f"F{r}"] = f"=E{r}*D{r}"
        ws[f"F{r}"].number_format = '#,##0'
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.font = font_sp
            cell.alignment = Alignment(horizontal="center")

    r_sum = dong_mau + so_sp
    r_vat = r_sum + 1
    ws[f"A{r_sum}"] = f'="Tổng: Mục ("&MIN(A{dong_mau}:A{dong_mau+so_sp-1})&" ~ "&MAX(A{dong_mau}:A{dong_mau+so_sp-1})&")"'
    ws.merge_cells(f"A{r_sum}:B{r_sum}")
    ws[f"D{r_sum}"] = f"=SUM(D{dong_mau}:D{dong_mau+so_sp-1})"
    ws[f"F{r_sum}"] = f"=SUM(F{dong_mau}:F{dong_mau+so_sp-1})"
    ws[f"F{r_sum}"].number_format = '#,##0'
    ws[f"A{r_vat}"] = "Thuế VAT"
    ws.merge_cells(f"A{r_vat}:B{r_vat}")
    ws[f"C{r_vat}"] = "%"
    ws[f"D{r_vat}"] = 5
    ws[f"F{r_vat}"] = f"=F{r_sum}*D{r_vat}%"
    ws[f"F{r_vat}"].number_format = '#,##0'
    r_total = r_vat + 1
    ws[f"A{r_total}"] = "Tổng"
    ws.merge_cells(f"A{r_total}:B{r_total}")
    ws[f"F{r_total}"] = f"=F{r_sum}+F{r_vat}"
    ws[f"F{r_total}"].number_format = '#,##0'
    r_bangchu = r_total + 1
    ws.merge_cells(f"A{r_bangchu}:F{r_bangchu}")
    ws[f"A{r_bangchu}"].value = f'= "Bằng chữ: " & VND(F{r_total})'
    ws[f"A{r_bangchu}"].alignment = Alignment(horizontal="center")
    ws[f"A{r_bangchu}"].font = font_bold
    for r in [r_sum, r_vat, r_total]:
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = border_all
            cell.font = font_bold if r == r_total else font_sp
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
