from fastapi import APIRouter, UploadFile, File, Form
from fastapi.responses import StreamingResponse, JSONResponse
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook, Workbook
from copy import copy
from datetime import datetime
from openpyxl.styles import Alignment, Font, Border, Side
import os
import base64
import sendgrid
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
from huggingface_hub import upload_file
import tempfile

router = APIRouter()

# ==== Tạo số báo giá ====
def tao_so_bao_gia():
    file = "counter.txt"
    if os.path.exists(file):
        with open(file, "r") as f:
            so = int(f.read().strip()) + 1
    else:
        so = 47
    with open(file, "w") as f:
        f.write(str(so))
    return f"2025/ĐL6/BG.{so:03d}"

# ==== Sao chép định dạng dòng ====
def sao_chep_dong(ws, src_row, dest_row):
    for col in range(1, ws.max_column + 1):
        c1 = ws.cell(row=src_row, column=col)
        c2 = ws.cell(row=dest_row, column=col)
        c2.value = c1.value
        if c1.has_style:
            c2._style = copy(c1._style)

# ==== Sinh file báo giá Excel ====
def dien_file_baogia(mau, ten_khach, data, so_bg, vanchuyen):
    wb = load_workbook(mau)
    ws = wb.active

    ws["A4"] = f"Số báo giá: {so_bg}"
    ws["A6"] = f"Kính gửi: {ten_khach}"
    ws["C4"] = f"Hà Nội, ngày {datetime.now().day} tháng {datetime.now().month} năm {datetime.now().year}"

    dong_mau = 9
    so_sp = len(data)
    for _ in range(6):
        ws.delete_rows(dong_mau)
    ws.insert_rows(dong_mau, so_sp + 5)

    font_sp = Font(name='Times New Roman', size=13)
    font_bold = Font(name='Times New Roman', size=13, bold=True)
    border_all = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for i, (_, row) in enumerate(data.iterrows()):
        r = dong_mau + i
        sao_chep_dong(ws, dong_mau - 1, r)
        ws[f"A{r}"] = i + 1
        ws[f"B{r}"].value = str(row["Ten_thiet_bi"])
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        do_dai = len(ws[f"B{r}"].value)
        ws.row_dimensions[r].height = max(20, 15 * ((do_dai // 30) + 1))
        ws[f"C{r}"] = "Cái"
        ws[f"D{r}"] = row["Số lượng"]
        ws[f"E{r}"] = row["Don_gia"]
        ws[f"E{r}"].number_format = '#,##0'
        ws[f"F{r}"] = f"=E{r}*D{r}"
        ws[f"F{r}"].number_format = '#,##0'
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.font = font_sp
            cell.alignment = Alignment(horizontal="center")

    r_sum = dong_mau + so_sp
    r_vat = r_sum + 1
    ws[f"A{r_sum}"] = f'="Tổng: Mục ("&MIN(A{dong_mau}:A{dong_mau+so_sp-1})&" ~ "&MAX(A{dong_mau}:A{dong_mau+so_sp-1})&")"'
    ws.merge_cells(f"A{r_sum}:B{r_sum}")
    ws[f"D{r_sum}"] = f"=SUM(D{dong_mau}:D{dong_mau+so_sp-1})"
    ws[f"F{r_sum}"] = f"=SUM(F{dong_mau}:F{dong_mau+so_sp-1})"
    ws[f"F{r_sum}"].number_format = '#,##0'

    ws[f"A{r_vat}"] = "Thuế VAT"
    ws.merge_cells(f"A{r_vat}:B{r_vat}")
    ws[f"C{r_vat}"] = "%"
    ws[f"D{r_vat}"] = 5
    ws[f"F{r_vat}"] = f"=F{r_sum}*D{r_vat}%"
    ws[f"F{r_vat}"].number_format = '#,##0'

    r_total = r_vat + 1
    ws[f"A{r_total}"] = "Tổng"
    ws.merge_cells(f"A{r_total}:B{r_total}")
    ws[f"F{r_total}"] = f"=F{r_sum}+F{r_vat}"
    ws[f"F{r_total}"].number_format = '#,##0'

    r_bangchu = r_total + 1
    ws.merge_cells(f"A{r_bangchu}:F{r_bangchu}")
    ws[f"A{r_bangchu}"].value = f'= "Bằng chữ: " & VND(F{r_total})'
    ws[f"A{r_bangchu}"].alignment = Alignment(horizontal="center")
    ws[f"A{r_bangchu}"].font = font_bold

    for r in [r_sum, r_vat, r_total]:
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = border_all
            cell.font = font_bold if r == r_total else font_sp

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ==== Gửi mail báo giá bằng SendGrid ====
def gui_mail_sendgrid(file, email_nguoinhan, sobg, ten_khach):
    try:
        from_email = "doluong6.quatest1@gmail.com"
        api_key = os.environ["SENDGRID_API_KEY"]
        sg = sendgrid.SendGridAPIClient(api_key)
        message = Mail(
            from_email=from_email,
            to_emails=email_nguoinhan,
            subject=f"Báo giá thiết bị - {sobg}",
            plain_text_content=f"""Kính gửi {ten_khach},\n\nCảm ơn quý khách đã quan tâm tới sản phẩm/dịch vụ của chúng tôi.\nVui lòng xem báo giá đính kèm.\n\nTrân trọng,\nTrung tâm Kỹ thuật Tiêu chuẩn Đo lường Chất lượng 1 – Đo Lường 6"""
        )

        file_bytes = file.read()
        encoded = base64.b64encode(file_bytes).decode()
        attachment = Attachment(
            FileContent(encoded),
            FileName(f"{sobg}.xlsx"),
            FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            Disposition("attachment")
        )
        message.attachment = attachment
        sg.send(message)
        return True
    except Exception as e:
        print("❌ Lỗi gửi email:", e)
        return False

# ==== Ghi log báo giá ====
def ghi_log(ten_khach, email, sobg, ds):
    path = "bao_gia_log.xlsx"
    df = pd.DataFrame(ds)
    df.insert(0, "Số BG", sobg)
    df.insert(1, "Khách", ten_khach)
    df.insert(2, "Email", email)
    df.insert(3, "Ngày", datetime.now().strftime("%Y-%m-%d %H:%M"))
    try:
        old = pd.read_excel(path)
        df = pd.concat([old, df], ignore_index=True)
    except:
        pass
    df.to_excel(path, index=False)

# ==== API tạo báo giá từ danh sách thủ công ====
@router.post("/tao_baogia_tu_danhsach")
async def tao_baogia_tu_danhsach(
    ten_khach: str = Form(...),
    email: str = Form(...),
    data: str = Form(...),
    vanchuyen: float = Form(0.0)
):
    try:
        df = pd.read_json(data)
        sobg = tao_so_bao_gia()
        file_xlsx = dien_file_baogia("Baogia.xlsx", ten_khach, df, sobg, vanchuyen)
        save_path = f"/tmp/{sobg.replace('/', '_')}.xlsx"
        with open(save_path, "wb") as f:
            f.write(file_xlsx.getbuffer())

        folder_name = ten_khach.replace("/", "-").replace("\\", "-").replace(" ", "_")
        upload_file(
            path_or_fileobj=save_path,
            path_in_repo=f"Khach_hang/{folder_name}/Bao_gia/{sobg.replace('/', '_')}.xlsx",
            repo_id="doluong6/QuanLy",
            repo_type="space",
            token=os.environ["HF_TOKEN"]
        )

        ghi_log(ten_khach, email, sobg, df.to_dict("records"))
        gui_mail_sendgrid(file_xlsx, email, sobg, ten_khach)

        return StreamingResponse(file_xlsx, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={sobg.replace('/', '_')}.xlsx"})

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

# ==== API tạo báo giá từ file phiếu yêu cầu ====
@router.post("/tao_baogia_tu_pyc")
async def tao_baogia_tu_pyc(file: UploadFile = File(...), vanchuyen: float = Form(0.0)):
    try:
        filename = file.filename
        ten_khach = filename.split("-")[1] if "-" in filename else "KhachHang"
        ten_khach_sach = ten_khach.strip().replace("/", "-").replace("\\", "-")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file.file.read())
            path = tmp.name

        wb = load_workbook(path, data_only=True)
        ws = wb.active

        data = []
        for row in ws.iter_rows(min_row=24):
            ten_tb = row[3].value
            so_luong = row[64].value
            don_gia = row[67].value
            if ten_tb and so_luong and don_gia:
                data.append({"Ten_thiet_bi": ten_tb, "Số lượng": so_luong, "Don_gia": don_gia})

        if not data:
            return JSONResponse(content={"error": "Không có thiết bị hợp lệ trong phiếu yêu cầu"}, status_code=400)

        df = pd.DataFrame(data)
        sobg = tao_so_bao_gia()
        file_xlsx = dien_file_baogia("Baogia.xlsx", ten_khach, df, sobg, vanchuyen=0)

        save_path = f"/tmp/{sobg.replace('/', '_')}.xlsx"
        with open(save_path, "wb") as f:
            f.write(file_xlsx.getbuffer())

        upload_file(
            path_or_fileobj=save_path,
            path_in_repo=f"Khach_hang/{ten_khach_sach}/Bao_gia/{sobg.replace('/', '_')}.xlsx",
            repo_id="doluong6/QuanLy",
            repo_type="space",
            token=os.environ["HF_TOKEN"]
        )

        email = "doluong6@quatest1.com.vn"
        ghi_log(ten_khach, email, sobg, data)
        gui_mail_sendgrid(file_xlsx, email, sobg, ten_khach)

        return StreamingResponse(file_xlsx, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={sobg.replace('/', '_')}.xlsx"})

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)
# ==== API: Danh sách báo giá gần đây ====
@router.get("/bao_gia_moi_nhat")
def bao_gia_moi_nhat():
    from huggingface_hub import list_repo_files
    try:
        all_files = list_repo_files("doluong6/QuanLy", repo_type="space", token=os.environ["HF_TOKEN"])
        bg_files = [f for f in all_files if f.startswith("Khach_hang/") and "/Bao_gia/" in f and f.endswith(".xlsx")]
        bg_files_sorted = sorted(bg_files, reverse=True)[:10]
        return {"danh_sach": bg_files_sorted}
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)
# ==== API: Trả bảng giá tổng hợp ====
@router.get("/bang_gia")
def bang_gia():
    try:
        import pandas as pd
        files = [
            "bang_gia_hoaly.xlsx",
            "bang_gia_dungtich_luuluong.xlsx",
            "bang_gia_nhiet_thietbiyte.xlsx",
            "bang_gia_co_dodai_apsuat.xlsx",
            "bang_gia_dien.xlsx"
        ]
        all_devices = []
        for file in files:
            if os.path.exists(file):
                df = pd.read_excel(file)
                for _, row in df.iterrows():
                    all_devices.append({
                        "ten": row["Ten_thiet_bi"],
                        "ma": file,
                        "Don_gia": row["Don_gia"]
                    })
        return {"bang_gia": all_devices}
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)
# ==== API: Tạo báo giá từ thiết bị thủ công hoặc bảng giá ====
@router.post("/tao_baogia_thucong")
async def tao_baogia_thucong(request: dict):
    try:
        ds = request.get("thiet_bi", [])
        ten_khach = request.get("ten_khach", "KhachHang")
        email = request.get("email", "doluong6@quatest1.com.vn")
        vanchuyen = float(request.get("vanchuyen", 0))

        df = pd.DataFrame(ds)
        sobg = tao_so_bao_gia()
        file_xlsx = dien_file_baogia("Baogia.xlsx", ten_khach, df, sobg, vanchuyen)

        safe_name = sobg.replace("/", "_")
        save_path = f"/tmp/{safe_name}.xlsx"
        with open(save_path, "wb") as f:
            f.write(file_xlsx.getbuffer())

        folder_name = ten_khach.replace("/", "-").replace("\\", "-").replace(" ", "_")
        upload_file(
            path_or_fileobj=save_path,
            path_in_repo=f"Khach_hang/{folder_name}/Bao_gia/{safe_name}.xlsx",
            repo_id="doluong6/QuanLy",
            repo_type="space",
            token=os.environ["HF_TOKEN"]
        )

        ghi_log(ten_khach, email, sobg, ds)
        gui_mail_sendgrid(file_xlsx, email, sobg, ten_khach)

        return {"ten_file": f"Khach_hang/{folder_name}/Bao_gia/{safe_name}.xlsx"}

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)
# ==== API: Gửi mail báo giá hàng loạt từ file Excel ====
@router.post("/gui_mail_hangloat")
async def gui_mail_hangloat(file: UploadFile = File(...)):
    import tempfile
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(file.file.read())
        tmp.close()

        df = pd.read_excel(tmp.name)
        log = []
        for _, row in df.iterrows():
            ten_khach = row["Tên khách"]
            email = row["Email"]
            ten_file = row["File_bao_gia"]  # ví dụ: Khach_hang/.../BG_043.xlsx

            from huggingface_hub import hf_hub_download
            local_path = hf_hub_download(
                repo_id="doluong6/QuanLy",
                filename=ten_file,
                repo_type="space",
                token=os.environ["HF_TOKEN"]
            )

            with open(local_path, "rb") as f:
                file_bytes = BytesIO(f.read())
                success = gui_mail_sendgrid(file_bytes, email, sobg=ten_file.split("/")[-1].replace(".xlsx", ""), ten_khach=ten_khach)
                log.append({"email": email, "status": "✅" if success else "❌"})

        return {"ket_qua": log}
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

