from fastapi import APIRouter, UploadFile, File, Form
from fastapi.responses import JSONResponse
from huggingface_hub import upload_file, delete_file, list_repo_files
from openpyxl import load_workbook
from io import BytesIO
import os
import re

router = APIRouter()
REPO_ID = "doluong6/QuanLy"

# ==== Làm sạch text để so khớp ====
def clean_text(text):
    return re.sub(r"\s+", " ", str(text).strip().lower())

# ==== Upload biên bản hiệu chuẩn ====
@router.post("/upload_bienban")
async def upload_bienban(file: UploadFile = File(...)):
    try:
        filename = file.filename
        tmp_path = f"/tmp/{filename}"
        with open(tmp_path, "wb") as f:
            f.write(file.file.read())

        wb = load_workbook(tmp_path, data_only=True)
        if "BB" not in wb.sheetnames:
            return JSONResponse(content={"error": "Không có sheet BB"}, status_code=400)
        ws = wb["BB"]

        co_so1 = clean_text(ws["I10"].value)
        co_so2 = clean_text(ws["I11"].value)
        ten_gop = f"{co_so1} {co_so2}".strip()

        # ==== Tìm thư mục PYC khớp với cơ sở sử dụng ====
        all_files = list_repo_files(repo_id=REPO_ID, repo_type="space", token=os.environ["HF_TOKEN"])
        pyc_files = [f for f in all_files if f.startswith("Khach_hang/") and "/Phieu_yeu_cau/" in f]

        thu_muc_goc = None
        for f in pyc_files:
            try:
                data = BytesIO()
                data.write(upload_file(f, repo_id=REPO_ID, repo_type="space", token=os.environ["HF_TOKEN"]).read())
                wb_pyc = load_workbook(data, data_only=True)
                ws_pyc = wb_pyc.active
                k11 = clean_text(ws_pyc["K11"].value)
                if ten_gop in k11:
                    thu_muc_goc = "/".join(f.split("/")[:2])
                    break
            except:
                continue

        if not thu_muc_goc:
            thu_muc_goc = "Khach_hang/Khac"

        path_in_repo = f"{thu_muc_goc}/Giay_chung_nhan/{filename}"

        upload_file(
            path_or_fileobj=tmp_path,
            path_in_repo=path_in_repo,
            repo_id=REPO_ID,
            repo_type="space",
            token=os.environ["HF_TOKEN"]
        )

        return JSONResponse(content={"message": f"✅ Đã upload biên bản vào {path_in_repo}"})

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ==== Xoá file biên bản hiệu chuẩn ====
@router.post("/delete_bienban")
async def delete_bienban(file_path: str = Form(...), password: str = Form(...)):
    try:
        if password != os.environ.get("USER_PASSWORD"):
            return JSONResponse(content={"error": "Sai mật khẩu hoặc không có quyền"}, status_code=403)

        delete_file(
            path_in_repo=file_path,
            repo_id=REPO_ID,
            repo_type="space",
            token=os.environ["HF_TOKEN"]
        )
        return JSONResponse(content={"message": f"✅ Đã xoá file {file_path}"})

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)
